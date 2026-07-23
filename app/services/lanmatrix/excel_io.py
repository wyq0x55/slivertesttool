"""Excel template / import / export for dynamic-field projects (FR-EXCEL-*).

Flask-independent (openpyxl only) so it is unit-testable in isolation.

* :func:`build_template` — a ``.xlsx`` template from a project's field specs,
  with a hidden field-key row, required markers, a "说明" sheet, and data-
  validation dropdowns for select fields.
* :func:`parse_import` — read a workbook into rows keyed by field_key (matching
  first on the hidden field-key row, then on display titles), sanitising each
  cell against formula injection and returning per-cell provenance for errors.
* :func:`build_export` — write rows to a ``.xlsx`` with formula-injection
  escaping on every text cell.

The heavy validation lives in :mod:`app.services.lanmatrix.validation`; this module only
handles the Excel <-> dict boundary.
"""

from __future__ import annotations

import io
from typing import Any, BinaryIO, Optional, Union

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import security

TEMPLATE_VERSION = "1.0"
DATA_SHEET = "TestItems"
INFO_SHEET = "说明"
# Row layout on the data sheet.
ROW_TITLE = 1        # human display titles (+ required marker)
ROW_FIELD_KEY = 2    # hidden machine field keys (authoritative for mapping)
ROW_DATA_START = 3


class ExcelIOError(Exception):
    """Raised when a workbook cannot be produced or parsed."""


def _as_seekable(source: Union[str, "BinaryIO"]) -> Union[str, io.BytesIO]:
    if isinstance(source, str) and not hasattr(source, "read"):
        return source
    if isinstance(source, io.BytesIO):
        source.seek(0)
        return source
    seekable = False
    try:
        seekable = bool(source.seekable())  # type: ignore[attr-defined]
    except Exception:  # noqa: BLE001
        seekable = False
    if seekable:
        try:
            source.seek(0)  # type: ignore[attr-defined]
            return source
        except Exception:  # noqa: BLE001
            pass
    try:
        try:
            source.seek(0)  # type: ignore[attr-defined]
        except Exception:  # noqa: BLE001
            pass
        data = source.read()  # type: ignore[attr-defined]
    except Exception as exc:  # noqa: BLE001
        raise ExcelIOError(f"could not read upload: {exc}") from exc
    if isinstance(data, str):
        data = data.encode("utf-8")
    return io.BytesIO(data)


# --------------------------------------------------------------------------- #
# Template
# --------------------------------------------------------------------------- #
def build_template(project: dict[str, Any], specs: list[dict[str, Any]]) -> Workbook:
    """Build an import template. ``specs`` is a list of field dicts."""
    wb = Workbook()
    ws = wb.active
    ws.title = DATA_SHEET

    active_specs = [s for s in specs if s.get("is_active", True)]
    for col, spec in enumerate(active_specs, start=1):
        letter = get_column_letter(col)
        title = spec.get("display_name") or spec["field_key"]
        if spec.get("is_required"):
            title += " *"
        ws.cell(ROW_TITLE, col, title)
        ws.cell(ROW_FIELD_KEY, col, spec["field_key"])
        ws.column_dimensions[letter].width = 22

        options = spec.get("options") or []
        if spec.get("data_type") in ("single_select", "multi_select") and options:
            formula = '"' + ",".join(str(o) for o in options)[:250] + '"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.add(f"{letter}{ROW_DATA_START}:{letter}1048576")
            ws.add_data_validation(dv)

    # Hide the machine field-key row.
    ws.row_dimensions[ROW_FIELD_KEY].hidden = True
    ws.freeze_panes = f"A{ROW_DATA_START}"

    info = wb.create_sheet(INFO_SHEET)
    info["A1"] = "项目编码"
    info["B1"] = project.get("code", "")
    info["A2"] = "模板版本"
    info["B2"] = TEMPLATE_VERSION
    info["A4"] = "字段标识"
    info["B4"] = "显示名称"
    info["C4"] = "必填"
    info["D4"] = "类型"
    info["E4"] = "说明"
    for i, spec in enumerate(active_specs, start=5):
        info.cell(i, 1, spec["field_key"])
        info.cell(i, 2, spec.get("display_name", ""))
        info.cell(i, 3, "是" if spec.get("is_required") else "否")
        info.cell(i, 4, spec.get("data_type", ""))
        info.cell(i, 5, spec.get("help_text", ""))
    return wb


# --------------------------------------------------------------------------- #
# Import
# --------------------------------------------------------------------------- #
def parse_import(
    source: Union[str, "BinaryIO"],
    specs: list[dict[str, Any]],
) -> dict[str, Any]:
    """Parse a workbook into rows keyed by field_key.

    Returns ``{"columns": {col_index: field_key}, "unmapped": [titles],
    "missing_required": [field_keys], "rows": [{"row": n, "values": {..},
    "raw_titles": {..}}]}``. Cells are sanitised against control chars; formula
    strings are not trusted.
    """
    try:
        wb = load_workbook(_as_seekable(source), data_only=True, read_only=True)
    except ExcelIOError:
        raise
    except Exception as exc:  # noqa: BLE001
        raise ExcelIOError(f"无法解析工作簿: {exc}") from exc

    if DATA_SHEET in wb.sheetnames:
        ws = wb[DATA_SHEET]
    else:
        ws = wb.worksheets[0]
    if ws.max_row is None or ws.max_row < 1:
        raise ExcelIOError("工作表为空")

    spec_by_key = {s["field_key"]: s for s in specs}
    title_to_key = {
        _norm_title(s.get("display_name") or s["field_key"]): s["field_key"]
        for s in specs
    }

    # Read the first two rows to detect field-key mapping.
    rows_iter = ws.iter_rows(values_only=True)
    header_rows = []
    for r in rows_iter:
        header_rows.append(r)
        if len(header_rows) >= 2:
            break

    title_row = header_rows[0] if header_rows else ()
    key_row = header_rows[1] if len(header_rows) > 1 else ()

    columns: dict[int, str] = {}
    unmapped: list[str] = []
    has_keyrow = any(isinstance(v, str) and v in spec_by_key for v in key_row)
    data_start_index = 3 if has_keyrow else 2

    source_row = key_row if has_keyrow else title_row
    for idx, val in enumerate(source_row):
        if val is None:
            continue
        sval = str(val).strip()
        if has_keyrow and sval in spec_by_key:
            columns[idx] = sval
        elif not has_keyrow:
            key = title_to_key.get(_norm_title(sval))
            if key:
                columns[idx] = key
            else:
                unmapped.append(sval)
    if not has_keyrow:
        # Titles came from row 1; nothing else to add.
        pass

    mapped_keys = set(columns.values())
    missing_required = [
        s["field_key"] for s in specs
        if s.get("is_required") and not s.get("is_readonly")
        and s["field_key"] not in mapped_keys
    ]

    out_rows: list[dict[str, Any]] = []
    all_values = list(ws.iter_rows(min_row=data_start_index, values_only=True))
    for offset, row in enumerate(all_values):
        excel_row = data_start_index + offset
        values: dict[str, Any] = {}
        raw_titles: dict[str, str] = {}
        nonempty = False
        for idx, key in columns.items():
            cell = row[idx] if idx < len(row) else None
            if isinstance(cell, str):
                cell = security.sanitize_incoming(cell)
            if cell not in (None, ""):
                nonempty = True
            values[key] = cell
            raw_titles[key] = get_column_letter(idx + 1)
        if not nonempty:
            continue
        out_rows.append({"row": excel_row, "values": values, "cols": raw_titles})

    return {
        "columns": {i: k for i, k in columns.items()},
        "unmapped": unmapped,
        "missing_required": missing_required,
        "rows": out_rows,
        "template_version": TEMPLATE_VERSION,
    }


def _norm_title(value: Any) -> str:
    return str(value or "").strip().rstrip("*").strip().lower()


# --------------------------------------------------------------------------- #
# Export
# --------------------------------------------------------------------------- #
def build_export(
    project: dict[str, Any],
    specs: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    *,
    columns: Optional[list[str]] = None,
) -> Workbook:
    """Write rows to a workbook, escaping formula-like text (FR-EXCEL-007)."""
    wb = Workbook()
    ws = wb.active
    ws.title = DATA_SHEET

    spec_by_key = {s["field_key"]: s for s in specs}
    order = columns or [s["field_key"] for s in specs if s.get("is_active", True)]

    for col, key in enumerate(order, start=1):
        spec = spec_by_key.get(key, {})
        ws.cell(ROW_TITLE, col, spec.get("display_name") or key)
        ws.cell(ROW_FIELD_KEY, col, key)
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.row_dimensions[ROW_FIELD_KEY].hidden = True
    ws.freeze_panes = f"A{ROW_DATA_START}"

    for r, record in enumerate(rows, start=ROW_DATA_START):
        for col, key in enumerate(order, start=1):
            value = record.get(key)
            if isinstance(value, list):
                value = ";".join(str(v) for v in value)
            value = security.escape_formula(value)
            ws.cell(r, col, value)

    info = wb.create_sheet(INFO_SHEET)
    info["A1"] = "项目编码"
    info["B1"] = project.get("code", "")
    info["A2"] = "导出时间"
    info["B2"] = project.get("_exported_at", "")
    info["A3"] = "数据版本"
    info["B3"] = project.get("_data_version", "")
    return wb


def workbook_bytes(wb: Workbook) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
