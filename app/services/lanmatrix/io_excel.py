"""Excel -> 入出力 (I/O signal pool) conversion (openpyxl, Flask-independent).

The 入出力 workbook is a flat table, one row per reusable input/expected signal,
driven by a **header row**. The header cells name the columns using the field
keys (or common Japanese/Chinese aliases); order and position are not fixed.
Recognised columns::

    io_name   (名称 / 名前 / signal / name / io_name)
    io_path   (路径 / パス / path / io_path)
    io_note   (备考 / 備考 / note / io_note)

Each row defines a signal referenced from a step as ``io_name(io_path)`` in a
single cell. ``parse_workbook`` returns plain dicts keyed by field key; the
DB-facing mapping / uniqueness enforcement lives in ``libconst_bridge``. Free of
Flask / SQLAlchemy imports for isolated testing.
"""
from __future__ import annotations

from typing import Any, BinaryIO, Optional, Union

from openpyxl import Workbook, load_workbook

from . import matrix_excel as _mx

# field key -> set of accepted header labels (lower-cased, stripped). Exact
# (not substring) match, so "データ名" never collides with "名".
_HEADER_ALIASES: dict[str, set[str]] = {
    "io_name": {"io_name", "名称", "名前", "信号名", "信号名称", "signal",
                "signal_name", "name"},
    "io_path": {"io_path", "路径", "パス", "path", "signal_path", "アクセスパス"},
    "io_note": {"io_note", "io_notes", "备考", "備考", "note", "notes", "说明"},
}
_ALL_KEYS = tuple(_HEADER_ALIASES.keys())
_MAX_HEADER_SCAN = 30


class IoExcelError(Exception):
    """Raised for a malformed 入出力 workbook."""


def parse_workbook(source: Union[str, BinaryIO],
                   *, source_filename: str = "") -> dict[str, Any]:
    try:
        # ``read_only=False``: this parser uses random ``ws.cell(r, c)`` access,
        # which is pathologically slow in openpyxl's read-only mode. The pool
        # table is small, so a normal load is both correct and fast.
        wb = load_workbook(_mx._as_seekable(source), data_only=True,
                           read_only=False)
    except Exception as exc:  # noqa: BLE001 - surface as a domain error
        raise IoExcelError(f"无法打开工作簿：{exc}") from exc

    items: list[dict[str, Any]] = []
    sheets: list[str] = []
    for ws in wb.worksheets:
        header = _find_header(ws)
        if header is None:
            continue
        header_row, colmap = header
        sheets.append(ws.title)
        for r in range(header_row + 1, ws.max_row + 1):
            row = {key: _cell_str(ws.cell(r, col).value)
                   for key, col in colmap.items()}
            # Skip blank / trailing rows: a real signal needs a name or a path.
            if not row.get("io_name") and not row.get("io_path"):
                continue
            items.append({key: row.get(key, "") for key in _ALL_KEYS})
    return {"items": items, "sheets": sheets}


# --------------------------------------------------------------------------- #
# Export (I/O items -> flat .xlsx). The header labels below are chosen from the
# recognised aliases so an exported file re-imports losslessly.
# --------------------------------------------------------------------------- #
_EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("io_name", "名称"),
    ("io_path", "路径"),
    ("io_note", "備考"),
]
_EXPORT_WIDTHS = [34, 44, 40]
_IO_SHEET_TITLE = "IO"


def build_workbook(matrix: dict[str, Any]) -> Workbook:
    """Rebuild an 入出力 ``.xlsx`` (flat table) from editor items.

    ``matrix`` is ``{"items": [ {io_name, io_path, io_note}, ... ]}``. Only the
    "No." index column is synthetic; every recognised field maps to a header the
    importer understands, so the round trip is lossless.
    """
    from openpyxl.utils import get_column_letter

    items = matrix.get("items") or []
    wb = Workbook()
    ws = wb.active
    ws.title = matrix.get("sheet_title") or _IO_SHEET_TITLE

    ws.append(["No."] + [jp for _key, jp in _EXPORT_COLUMNS])
    keys = [key for key, _jp in _EXPORT_COLUMNS]
    n = 0
    for it in items:
        if _is_blank_io(it):
            continue
        n += 1
        ws.append([n] + [_out(it.get(k)) for k in keys])

    ws.column_dimensions["A"].width = 6
    for i, width in enumerate(_EXPORT_WIDTHS):
        ws.column_dimensions[get_column_letter(i + 2)].width = width
    ws.freeze_panes = "A2"
    return wb


def _is_blank_io(item: dict) -> bool:
    if not isinstance(item, dict):
        return True
    return not (item.get("io_name") or item.get("io_path"))


def _out(val: Any) -> Any:
    return None if val in (None, "") else val


def _find_header(ws) -> Optional[tuple[int, dict[str, int]]]:
    """Locate the header row and map each recognised field key to its column."""
    for r in range(1, min(_MAX_HEADER_SCAN, ws.max_row) + 1):
        colmap: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            key = _match_header(ws.cell(r, c).value)
            if key and key not in colmap:
                colmap[key] = c
        # A valid 入出力 header must at least name the signal or its path.
        if "io_name" in colmap or "io_path" in colmap:
            return r, colmap
    return None


def _match_header(val: Any) -> Optional[str]:
    s = _cell_str(val).lower()
    if not s:
        return None
    for key, aliases in _HEADER_ALIASES.items():
        if s in aliases:
            return key
    return None


def _cell_str(val: Any) -> str:
    return "" if val is None else str(val).strip()
