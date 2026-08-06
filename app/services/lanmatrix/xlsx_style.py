"""Shared cell styling for every Lan-Matrix ``.xlsx`` export.

Historically each exporter wrote bare cells (no font, no borders, no fill), so
downloaded workbooks looked flat and unlike the hand-authored reference files.
This module centralises the exact visual style those reference workbooks use so
all export paths share a single definition.

Reference spec (derived from ``LibFunc``'s ``状態遷移`` sheet):

* **Font** — everywhere: ``Meiryo UI`` 8pt, regular (not bold), black.
* **Fill** (header rows only): light green ``FFCCFFCC`` on the fixed columns
  (手順番号 .. 引数) and the 入力値 signal columns; light blue ``FFCCECFF`` on the
  期待値 signal columns and the 確認タイミング column. Data rows: no fill.
* **Borders** — thin black. The three header rows of the fixed columns
  (手順番号 .. 引数) are framed as **one block** (top on the first row, bottom on
  the last, left/right on all, no internal horizontal lines) so they read as a
  single tall cell even though they are not merged. The signal-column header
  cells and every data cell get a full four-sided box.
* **Alignment / wrap** — per column: the fixed label columns are top-aligned;
  手順番号 data is left/middle; 手順目的 / 操作手順 wrap; サブルーチン / 引数 are
  top-aligned; the signal value columns are centered and wrapped.
* **Column widths** — margin column 4.125, 手順番号 6.625, 手順目的 46.0,
  操作手順 40.625; every signal column (入力値 / 期待値 / 確認タイミング) is 22.
  サブルーチン / 引数 keep Excel's default so wrapped text stays compact.

Flask-independent (openpyxl only) so the exporter unit tests import it directly.
"""
from __future__ import annotations

from typing import Iterable

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Header fills. Green = input side, blue = expected side.
HEADER_FILL_RGB = "FFCCFFCC"          # light green — fixed columns + 入力値
HEADER_EXPECT_FILL_RGB = "FFCCECFF"   # light blue  — 期待値 + 確認タイミング

HEADER_FILL = PatternFill(patternType="solid", fgColor=HEADER_FILL_RGB)
HEADER_EXPECT_FILL = PatternFill(patternType="solid",
                                 fgColor=HEADER_EXPECT_FILL_RGB)

# Every table cell uses the reference font.
CELL_FONT = Font(name="Meiryo UI", size=8, bold=False)

_THIN = Side(style="thin", color="FF000000")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Step-table column offsets relative to the first column (手順番号).
_OFF_NO, _OFF_PURPOSE, _OFF_OP, _OFF_SUB, _OFF_ARG = 0, 1, 2, 3, 4
_FIRST_SIGNAL_OFFSET = 5  # 入力値 / 期待値 / 確認タイミング begin here

# Column widths (reference ``状態遷移`` sheet). Only the columns the reference
# gives an explicit custom width are set; the others keep Excel's default so long
# サブルーチン / 引数 / signal values wrap instead of stretching the sheet — exactly
# how the reference lays them out.
_STEP_SPACER_WIDTH = 4.125          # the thin margin column left of 手順番号
_STEP_COL_WIDTHS = {                # offset from 手順番号 -> column width
    _OFF_NO: 6.625,       # 手順番号
    _OFF_PURPOSE: 46.0,   # 手順目的
    _OFF_OP: 40.625,      # 操作手順
}
# All signal columns (入力値 / 期待値 / 確認タイミング) share one width.
_STEP_SIGNAL_WIDTH = 22


# --------------------------------------------------------------------------- #
# Alignment helpers
# --------------------------------------------------------------------------- #
_SIGNAL_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _header_align(offset: int) -> Alignment:
    if offset >= _FIRST_SIGNAL_OFFSET:
        return _SIGNAL_ALIGN
    # Fixed label columns: top-aligned; サブルーチン wraps (longest label).
    return Alignment(vertical="top", wrap_text=(offset == _OFF_SUB))


def _data_align(offset: int) -> Alignment:
    if offset >= _FIRST_SIGNAL_OFFSET:
        return _SIGNAL_ALIGN
    if offset == _OFF_NO:
        return Alignment(horizontal="left", vertical="center")
    if offset == _OFF_PURPOSE:
        return Alignment(vertical="center", wrap_text=True)
    if offset == _OFF_OP:
        return Alignment(horizontal="left", vertical="center", wrap_text=True)
    if offset == _OFF_SUB:
        return Alignment(vertical="top", wrap_text=True)
    return Alignment(vertical="top")  # 引数


# Description ("说明") block above each step table: the test-id / title row and
# the metadata label rows. Reference: Meiryo UI 8, left / vertical-center; the
# test-id cell is bold, the title (テスト名) cell wraps; no borders, no fill.
_META_FONT_BOLD = Font(name="Meiryo UI", size=8, bold=True)
_META_ALIGN = Alignment(horizontal="left", vertical="center")
_META_TITLE_ALIGN = Alignment(horizontal="left", vertical="center",
                              wrap_text=True)


def style_detail_meta(
    ws,
    id_row: int,
    meta_first_row: int,
    meta_last_row: int,
    *,
    id_col: int,
    title_col: int,
    label_col: int,
    value_col: int,
) -> None:
    """Style the description block that sits above a step table.

    ``id_row`` carries the test-id / lib-func in ``id_col`` (bold) and the test
    name / title in ``title_col`` (wrapped). Rows ``meta_first_row`` ..
    ``meta_last_row`` are label/value pairs (``label_col`` / ``value_col``).
    All cells use the reference font; none get borders or fill.
    """
    idc = ws.cell(id_row, id_col)
    idc.font = _META_FONT_BOLD
    idc.alignment = _META_ALIGN
    tc = ws.cell(id_row, title_col)
    tc.font = CELL_FONT
    tc.alignment = _META_TITLE_ALIGN
    for r in range(meta_first_row, meta_last_row + 1):
        for c in (label_col, value_col):
            cell = ws.cell(r, c)
            cell.font = CELL_FONT
            cell.alignment = _META_ALIGN


def _group_border(row: int, top_row: int, bottom_row: int) -> Border:
    """Outer frame for one row of the fixed-column header block."""
    return Border(
        left=_THIN, right=_THIN,
        top=_THIN if row == top_row else None,
        bottom=_THIN if row == bottom_row else None,
    )


# --------------------------------------------------------------------------- #
# Flat tables (const / io / dynamic-field export): one green header row.
# --------------------------------------------------------------------------- #
_FLAT_HEADER_ALIGN = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
_FLAT_DATA_ALIGN = Alignment(vertical="center", wrap_text=True)


def style_region(
    ws,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    *,
    header_rows: Iterable[int] = (),
) -> None:
    """Style a flat table: Meiryo UI font, thin box on every cell, green fill +
    centered header on ``header_rows``. No-ops on an empty region.
    """
    if max_row < min_row or max_col < min_col:
        return
    header_set = set(header_rows)
    for r in range(min_row, max_row + 1):
        is_header = r in header_set
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
            if is_header:
                cell.fill = HEADER_FILL
                cell.alignment = _FLAT_HEADER_ALIGN
            else:
                cell.alignment = _FLAT_DATA_ALIGN


# --------------------------------------------------------------------------- #
# Procedure step tables (test-matrix / lib-func detail blocks).
# --------------------------------------------------------------------------- #
def style_step_table(
    ws,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    *,
    header_rows: Iterable[int],
    expect_start_col: int,
) -> None:
    """Style a procedure step table to match the reference workbook.

    ``min_col`` is the 手順番号 column; the layout after it is fixed
    (手順目的/操作手順/サブルーチン/引数, then the 入力値/期待値 signal columns and
    finally 確認タイミング at ``max_col``). ``expect_start_col`` is the first
    期待値 column (``COL_SIGNAL_START + n_in``); columns from there on are filled
    blue, the rest of the header green.

    Borders: the fixed-column header cells (手順番号 .. 引数) are framed as a
    single block across the header rows (no internal horizontal lines); signal
    header cells and all data cells get a full four-sided box.
    """
    if max_row < min_row or max_col < min_col:
        return
    header_set = set(header_rows)
    hrows = sorted(r for r in header_set if min_row <= r <= max_row)
    top_hr = hrows[0] if hrows else min_row
    bot_hr = hrows[-1] if hrows else min_row
    signal_start = min_col + _FIRST_SIGNAL_OFFSET

    for r in range(min_row, max_row + 1):
        is_header = r in header_set
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            cell.font = CELL_FONT
            offset = c - min_col
            if is_header:
                cell.fill = (HEADER_EXPECT_FILL if c >= expect_start_col
                             else HEADER_FILL)
                cell.alignment = _header_align(offset)
                # Fixed label columns share one outer frame; signal columns box.
                if c < signal_start:
                    cell.border = _group_border(r, top_hr, bot_hr)
                else:
                    cell.border = THIN_BORDER
            else:
                cell.alignment = _data_align(offset)
                cell.border = THIN_BORDER

    # Column widths — mirror the reference's explicit custom widths. Setting them
    # here (once per block) is idempotent: repeated blocks on one sheet resolve to
    # the same per-column value.
    if min_col - 1 >= 1:
        ws.column_dimensions[get_column_letter(min_col - 1)].width = \
            _STEP_SPACER_WIDTH
    for off, width in _STEP_COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(min_col + off)].width = width
    for c in range(signal_start, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = _STEP_SIGNAL_WIDTH
