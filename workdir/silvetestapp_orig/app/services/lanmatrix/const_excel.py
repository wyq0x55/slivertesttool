"""Excel -> Const conversion (openpyxl, Flask-independent).

The constant-definition workbook ("Const") is a flat table, one row per
constant, driven by a **header row**. The header cells name the columns using
the field keys (or common Japanese/Chinese aliases); order and position are not
fixed. Recognised columns (matches the real "MCU" constant workbook layout)::

    const_name      (识别子名 / 識別子名 / identifier / const_name)
    const_jname     (和名 / const_jname)
    const_value     (値 / value / const_value)
    const_class1    (分类1 / 分類1 / const_class1)
    const_class2    (分类2 / 分類2 / const_class2)
    const_dataname  (数据名称 / データ名称 / const_dataname)
    const_note      (备考 / 備考 / note / const_note)

Column order and position are not fixed — the header row is located by matching
these labels. ``parse_workbook`` returns plain dicts keyed by field key; the
DB-facing mapping lives in ``libconst_bridge``. Free of Flask / SQLAlchemy
imports for isolated testing.
"""
from __future__ import annotations

from typing import Any, BinaryIO, Optional, Union

from openpyxl import load_workbook

from . import matrix_excel as _mx

# field key -> set of accepted header labels (lower-cased, stripped). Exact
# (not substring) match, so "データ名称" never collides with "名称".
_HEADER_ALIASES: dict[str, set[str]] = {
    "const_name": {"const_name", "识别子名", "識別子名", "识别子", "識別子",
                   "identifier", "ident", "identifier_name"},
    "const_jname": {"const_jname", "和名", "常量名", "常数名", "jname"},
    "const_value": {"const_value", "值", "値", "value"},
    "const_class1": {"const_class1", "分类1", "分類1", "class1", "category1"},
    "const_class2": {"const_class2", "分类2", "分類2", "class2", "category2"},
    "const_dataname": {"const_dataname", "数据名称", "データ名称", "データ名",
                       "dataname", "data_name", "数据名"},
    "const_note": {"const_note", "const_notes", "备考", "備考", "note", "notes",
                   "说明"},
}
_ALL_KEYS = tuple(_HEADER_ALIASES.keys())
_MAX_HEADER_SCAN = 30


class ConstExcelError(Exception):
    """Raised for a malformed Const workbook."""


def parse_workbook(source: Union[str, BinaryIO],
                   *, source_filename: str = "") -> dict[str, Any]:
    try:
        # ``read_only=False``: this parser uses random ``ws.cell(r, c)`` access,
        # which is pathologically slow in openpyxl's read-only mode. The Const
        # table is small, so a normal load is both correct and fast.
        # ``_as_seekable`` buffers non-seekable upload streams (Werkzeug's
        # ``SpooledTemporaryFile``) so the underlying zipfile reader works.
        wb = load_workbook(_mx._as_seekable(source), data_only=True,
                           read_only=False)
    except Exception as exc:  # noqa: BLE001 - surface as a domain error
        raise ConstExcelError(f"无法打开工作簿：{exc}") from exc

    items: list[dict[str, Any]] = []
    sheets: list[str] = []
    for ws in wb.worksheets:
        header = _find_header(ws)
        if header is None:
            continue
        header_row, colmap = header
        sheets.append(ws.title)
        for r in range(header_row + 1, ws.max_row + 1):
            row = {key: _cell_str(ws.cell(r, col).value)
                   for key, col in colmap.items()}
            # Skip blank / trailing rows: a real constant needs a name or a value.
            if not row.get("const_name") and not row.get("const_jname") \
                    and not row.get("const_value"):
                continue
            items.append({key: row.get(key, "") for key in _ALL_KEYS})
    return {"items": items, "sheets": sheets}


def _find_header(ws) -> Optional[tuple[int, dict[str, int]]]:
    """Locate the header row and map each recognised field key to its column."""
    for r in range(1, min(_MAX_HEADER_SCAN, ws.max_row) + 1):
        colmap: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            key = _match_header(ws.cell(r, c).value)
            if key and key not in colmap:
                colmap[key] = c
        # A valid Const header must at least name the constant identifier,
        # its Japanese name, or its value column.
        if "const_name" in colmap or "const_jname" in colmap \
                or "const_value" in colmap:
            return r, colmap
    return None


def _match_header(val: Any) -> Optional[str]:
    s = _cell_str(val).lower()
    if not s:
        return None
    for key, aliases in _HEADER_ALIASES.items():
        if s in aliases:
            return key
    return None


def _cell_str(val: Any) -> str:
    return "" if val is None else str(val).strip()
