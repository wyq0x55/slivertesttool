"""Tests for the Flask-independent Excel codec (:mod:`app.services.lanmatrix.matrix_excel`).

These tests use only :mod:`openpyxl` + the stdlib ``unittest`` runner, so they
run without Flask / SQLAlchemy installed. They cover:

* parsing the reference workbook,
* a lossless parse -> build -> reparse round trip,
* export behaviour for summary-only items (no detail block), and
* the error path for a workbook without the ``DB`` summary table.

Run standalone::

    python -m unittest tests.test_matrix_excel
"""

from __future__ import annotations

import importlib.util
import io
import os
import unittest

from openpyxl import Workbook

_HERE = os.path.dirname(os.path.abspath(__file__))
_MODULE_PATH = os.path.join(
    _HERE, "..", "app", "services", "lanmatrix", "matrix_excel.py")
# Candidate locations for the reference workbook (repo copy or uploads).
_REFERENCE_CANDIDATES = [
    os.path.join(_HERE, "data", "VHILS1_RWS_MCU_Qualification_Test_SYS.xlsx"),
    "/app/uploads/VHILS1_RWS_MCU_Qualification_Test_SYS.xlsx",
]


def _load_module():
    spec = importlib.util.spec_from_file_location("matrix_excel", _MODULE_PATH)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


mx = _load_module()


def _reference_path():
    for path in _REFERENCE_CANDIDATES:
        if os.path.exists(path):
            return path
    return None


class RoundTripTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.ref = _reference_path()

    def setUp(self):
        if not self.ref:
            self.skipTest("reference workbook not available")

    def test_parse_reference(self):
        parsed = mx.parse_workbook(self.ref, source_filename=os.path.basename(self.ref))
        self.assertTrue(parsed["items"], "expected at least one item")
        self.assertTrue(parsed["id_prefix"], "expected a detected id prefix")
        for item in parsed["items"]:
            self.assertIn("category", item)
            self.assertIn("test_no", item)
            self.assertIn("steps", item)

    def test_round_trip_lossless(self):
        parsed = mx.parse_workbook(self.ref, source_filename=os.path.basename(self.ref))
        wb = mx.build_workbook(parsed)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        reparsed = mx.parse_workbook(buffer, source_filename=os.path.basename(self.ref))

        self.assertEqual(len(parsed["items"]), len(reparsed["items"]))
        self.assertEqual(parsed["id_prefix"], reparsed["id_prefix"])

        for a, b in zip(parsed["items"], reparsed["items"]):
            for key in ("category", "test_no", "test_name", "priority", "result"):
                self.assertEqual(a.get(key), b.get(key), f"mismatch on {key}")
            self.assertEqual(a.get("steps"), b.get("steps"), "steps mismatch")


class BuildTest(unittest.TestCase):
    def _summary_only_matrix(self):
        return {
            "name": "unit",
            "source_filename": "unit.xlsx",
            "summary_sheet": mx.DEFAULT_SUMMARY_SHEET,
            "id_prefix": "ID;;PFX-",
            "items": [
                {
                    "category": 1,
                    "test_no": 1,
                    "test_name": "summary only",
                    "priority": "\u4e0d\u8981",
                    "result": "-",
                    "steps": {"input_signals": [], "expected_signals": [], "steps": []},
                }
            ],
        }

    def test_summary_only_item_has_no_detail_sheet(self):
        wb = mx.build_workbook(self._summary_only_matrix())
        # Only the summary sheet exists; no numeric detail sheet was created.
        self.assertIn(mx.DEFAULT_SUMMARY_SHEET, wb.sheetnames)
        numeric = [s for s in wb.sheetnames if s.strip().isdigit()]
        self.assertEqual(numeric, [], "summary-only item must not produce a detail sheet")

    def test_build_reparse_summary_only(self):
        wb = mx.build_workbook(self._summary_only_matrix())
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        parsed = mx.parse_workbook(buffer, source_filename="unit.xlsx")
        self.assertEqual(len(parsed["items"]), 1)
        item = parsed["items"][0]
        self.assertEqual(item["category"], 1)
        self.assertEqual(item["test_no"], 1)
        self.assertEqual(item["steps"]["steps"], [])


class _NoSeekableStream:
    """A minimal file-like without ``seekable`` (mimics Werkzeug uploads)."""

    def __init__(self, data: bytes):
        self._buf = io.BytesIO(data)

    def read(self, *a):
        return self._buf.read(*a)

    def seek(self, *a):
        return self._buf.seek(*a)


class UploadStreamTest(unittest.TestCase):
    def setUp(self):
        self.ref = _reference_path()
        if not self.ref:
            self.skipTest("reference workbook not available")

    def test_parse_stream_without_seekable(self):
        with open(self.ref, "rb") as fh:
            data = fh.read()
        stream = _NoSeekableStream(data)
        parsed = mx.parse_workbook(stream, source_filename="x.xlsx")
        self.assertTrue(parsed["items"])


class ErrorTest(unittest.TestCase):
    def test_missing_db_table_raises(self):
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.active["A1"] = "not a matrix"
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        with self.assertRaises(mx.MatrixExcelError):
            mx.parse_workbook(buffer, source_filename="bad.xlsx")


if __name__ == "__main__":
    unittest.main()
