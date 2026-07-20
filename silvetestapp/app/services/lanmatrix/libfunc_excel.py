"""Excel -> Lib(Func) conversion (openpyxl, Flask-independent).

The function-library workbook ("Lib") is **block-structured**, not a flat table.
Each worksheet holds one or more *function blocks*, and every block looks like::

    | No. | <lib_func>       |    | <lib_name / 題名>            |   <- block header
    |     | 目的             | :<purpose>                          |
    |     | 引数             | :<arguments>                        |
    |     | 仮引数           | :<formal parameters>                |
    |     | 備考             | :<note>                             |
    |                                                                |   <- blank
    | 手順番号|手順目的|操作手順|サブルーチン|引数|入力値…|期待値…|確認タイミング |  <- step header
    |         |        |        |            |    |<signal display names>       |
    |         |        |        |            |    |<signal identifiers>         |
    |   1     | …      | …      | …          | …  | …                           |  <- step rows
    |   2     | …                                                             |
    (blank row separates the next block)

The step table geometry (``手順番号 .. 確認タイミング`` header + two signal-name
rows + step rows) is **identical** to the Test-Matrix detail sheets, so this
parser reuses :func:`matrix_excel.parse_step_table` verbatim to build ``lib_stb``
— the same JSON document (input/expected signals + step rows) the step editor
uses.

Sheets are classified as *initialization* (``isinit=True``) or not: if any
worksheet's title looks like an init sheet (``初期化`` / ``init`` …) the title
decides; otherwise the **first** worksheet is treated as the init sheet (per the
convention that the first Lib sheet holds initialization functions).

``parse_workbook`` returns::

    {"items": [ {isinit, lib_func, lib_name, lib_value, lib_arg,
                 lib_para, lib_note, lib_stb}, ... ], "sheets": [name, ...]}

``lib_stb`` is a dict; ``libconst_bridge.map_lib_item`` serialises it to JSON.
Free of Flask / SQLAlchemy imports so it can be unit-tested in isolation.
"""
from __future__ import annotations

from typing import Any, BinaryIO, Optional, Union

from openpyxl import load_workbook

from . import matrix_excel as _mx

_FIELD_KEYS = ("isinit", "lib_func", "lib_name", "lib_value", "lib_arg",
               "lib_para", "lib_note", "lib_stb")

# Block metadata labels (column B) -> field key. The value sits in column C.
_META_LABELS: dict[str, str] = {
    "目的": "lib_value",
    "引数": "lib_arg",
    "仮引数": "lib_para",
    "備考": "lib_note",
    "备考": "lib_note",
}
_META_LABEL_SET = frozenset(_META_LABELS.keys())

_STEP_HEADER_LABEL = "手順番号"

# Column positions (1-based), shared with the Test-Matrix detail sheets.
_COL_NO = _mx.COL_STEP_NO        # B — block No. / step number
_COL_META_VAL = 3                # C — metadata value cell
_COL_TITLE = 4                   # D — block 題名 (lib_name)

# Sheet-title tokens that mark an "initialization" worksheet.
_INIT_TOKENS = ("初期化", "初期", "初始化", "init")


class LibExcelError(Exception):
    """Raised for a malformed Lib(Func) workbook."""


def parse_workbook(source: Union[str, BinaryIO],
                   *, source_filename: str = "") -> dict[str, Any]:
    try:
        # Random ``ws.cell(r, c)`` access is pathologically slow in read-only
        # mode; the Lib workbook is small so a normal load is correct and fast.
        # ``_as_seekable`` buffers non-seekable upload streams (Werkzeug's
        # ``SpooledTemporaryFile``) so the underlying zipfile reader works.
        wb = load_workbook(_mx._as_seekable(source), data_only=True,
                           read_only=False)
    except Exception as exc:  # noqa: BLE001 - surface as a domain error
        raise LibExcelError(f"无法打开工作簿：{exc}") from exc

    worksheets = list(wb.worksheets)
    init_flags = _classify_init_sheets(worksheets)

    items: list[dict[str, Any]] = []
    sheets: list[str] = []
    for ws, is_init in zip(worksheets, init_flags):
        headers = _find_block_headers(ws)
        if headers:
            sheets.append(ws.title)
            for hr in headers:
                items.append(_parse_block(ws, hr, is_init))
            continue
        # Fallback: a flat header-driven table (legacy / hand-authored files).
        flat = _parse_flat_sheet(ws, is_init)
        if flat:
            sheets.append(ws.title)
            items.extend(flat)
    return {"items": items, "sheets": sheets}


# --------------------------------------------------------------------------- #
# Sheet classification
# --------------------------------------------------------------------------- #
def _is_init_title(title: str) -> bool:
    t = (title or "").strip().lower()
    return any(tok.lower() in t for tok in _INIT_TOKENS)


def _classify_init_sheets(worksheets: list) -> list[bool]:
    """Return an ``isinit`` flag per worksheet.

    If any sheet title looks like an init sheet, titles decide. Otherwise the
    first worksheet is treated as the initialization sheet (the first Lib sheet
    holds initialization functions).
    """
    titled = [_is_init_title(ws.title) for ws in worksheets]
    if any(titled):
        return titled
    return [i == 0 for i in range(len(worksheets))]


# --------------------------------------------------------------------------- #
# Block parsing
# --------------------------------------------------------------------------- #
def _find_block_headers(ws) -> list[int]:
    """Row indices of every function-block header on the sheet."""
    return [r for r in range(1, ws.max_row + 1) if _is_block_header(ws, r)]


def _is_block_header(ws, r: int) -> bool:
    """A block header has a No. in col A and a function id in col B (which is
    neither a metadata label nor the step-table header)."""
    no = ws.cell(r, 1).value
    func = _cell_str(ws.cell(r, _COL_NO).value)
    if no is None or not func:
        return False
    if func in _META_LABEL_SET or func == _STEP_HEADER_LABEL:
        return False
    return _looks_numeric(no)


def _parse_block(ws, header_row: int, is_init: bool) -> dict[str, Any]:
    item: dict[str, Any] = {key: "" for key in _FIELD_KEYS}
    item["isinit"] = bool(is_init)
    item["lib_func"] = _cell_str(ws.cell(header_row, _COL_NO).value)
    item["lib_name"] = _cell_str(ws.cell(header_row, _COL_TITLE).value)

    # Metadata rows: from just after the header up to the step-table header
    # (手順番号) or the next block, whichever comes first.
    r = header_row + 1
    while r <= ws.max_row:
        label = _cell_str(ws.cell(r, _COL_NO).value)
        if label == _STEP_HEADER_LABEL or _is_block_header(ws, r):
            break
        key = _META_LABELS.get(label)
        if key:
            item[key] = _strip_leading_colon(_cell_str(ws.cell(r, _COL_META_VAL).value))
        r += 1

    # Step table (手順) — reuse the shared Test-Matrix procedure-table parser.
    step_header = _mx._find_step_header(ws, header_row)
    if step_header is not None:
        try:
            in_sig, exp_sig, steps = _mx.parse_step_table(
                ws, step_header, stop_on_block=_is_block_header)
        except Exception:  # noqa: BLE001 - never let one bad block abort import
            in_sig, exp_sig, steps = [], [], []
    else:
        in_sig, exp_sig, steps = [], [], []
    item["lib_stb"] = {
        "input_signals": in_sig,
        "expected_signals": exp_sig,
        "steps": steps,
    }
    return item


# --------------------------------------------------------------------------- #
# Flat-table fallback (legacy header-driven layout)
# --------------------------------------------------------------------------- #
_FLAT_ALIASES: dict[str, set[str]] = {
    "isinit": {"isinit", "is_init", "初始化", "初期化", "init"},
    "lib_func": {"lib_func", "函数标识", "func", "function", "関数名"},
    "lib_name": {"lib_name", "函数名称", "name", "名称", "題名"},
    "lib_value": {"lib_value", "目的", "值", "値", "value", "purpose"},
    "lib_arg": {"lib_arg", "引数", "argument", "arguments"},
    "lib_para": {"lib_para", "仮引数", "参数", "para", "parameters"},
    "lib_note": {"lib_note", "备考", "備考", "note", "notes", "说明"},
    "lib_stb": {"lib_stb", "测试手顺", "測試手順", "手順", "手顺", "stb", "steps"},
}
_FLAT_MAX_HEADER_SCAN = 30


def _parse_flat_sheet(ws, is_init: bool) -> list[dict[str, Any]]:
    header = _find_flat_header(ws)
    if header is None:
        return []
    header_row, colmap = header
    out: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {key: _cell_str(ws.cell(r, col).value) for key, col in colmap.items()}
        if not row.get("lib_func") and not row.get("lib_name"):
            continue
        item = {key: row.get(key, "") for key in _FIELD_KEYS}
        item["isinit"] = _coerce_flat_bool(row.get("isinit"), is_init)
        item["lib_stb"] = row.get("lib_stb", "")
        out.append(item)
    return out


def _coerce_flat_bool(raw: Optional[str], default: bool) -> bool:
    if raw is None or raw == "":
        return bool(default)
    return str(raw).strip().lower() in ("1", "true", "yes", "y", "是", "○", "init")


def _find_flat_header(ws) -> Optional[tuple[int, dict[str, int]]]:
    for r in range(1, min(_FLAT_MAX_HEADER_SCAN, ws.max_row) + 1):
        colmap: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            key = _match_flat_header(ws.cell(r, c).value)
            if key and key not in colmap:
                colmap[key] = c
        if "lib_func" in colmap or "lib_name" in colmap:
            return r, colmap
    return None


def _match_flat_header(val: Any) -> Optional[str]:
    s = _cell_str(val).lower()
    if not s:
        return None
    for key, aliases in _FLAT_ALIASES.items():
        if s in aliases:
            return key
    return None


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _looks_numeric(val: Any) -> bool:
    if isinstance(val, bool):
        return False
    if isinstance(val, (int, float)):
        return True
    try:
        int(str(val).strip())
        return True
    except (TypeError, ValueError):
        return False


def _strip_leading_colon(text: str) -> str:
    t = text.lstrip()
    if t[:1] in (":", "："):
        t = t[1:]
    return t.strip()


def _cell_str(val: Any) -> str:
    return "" if val is None else str(val).strip()
