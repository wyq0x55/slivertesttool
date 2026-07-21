"""Excel <-> test-matrix conversion (openpyxl, Flask-independent).

This module is the single source of truth for the on-disk Excel layout used by
the "LAN Test Matrix" feature. It is deliberately free of any Flask / SQLAlchemy
imports so it can be unit-tested and reused in isolation.

The source workbook (``VHILS1_RWS_MCU_Qualification_Test_SYS.xlsx``) has two
synchronised representations of the same data:

1. A **summary sheet** (default name ``4.TestRequirement``) holding an Excel
   Table named ``DB``. One row per test item, 22 columns. ``テストID`` and
   ``ログ`` are calculated columns.
2. One **category detail sheet** per ``テスト区分`` (named by the category
   number, e.g. ``1``, ``2``, ``3``). Each test item is rendered as a block:
   a header row (``テストID`` + a ``テスト名`` VLOOKUP), 13 metadata label rows
   (VLOOKUP into ``DB``), then a procedure table (``手順番号`` .. ``確認タイミング``)
   with two signal-name header rows and the step rows.

``parse_workbook`` reads a workbook into plain dicts; ``build_workbook``
regenerates a byte-compatible workbook from those dicts. Round-tripping through
these two functions is lossless for the business data.
"""

from __future__ import annotations

import datetime as _dt
import io
import re
from typing import Any, BinaryIO, Optional, Union

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Config is a pure, Flask-independent settings class (values from ``.env``), so
# importing it here keeps this module free of Flask/SQLAlchemy dependencies.
try:
    # Normal case: imported inside the full ``app`` package.
    from ...config import Config as _Config
except ImportError:
    # Isolated loaders (the pure-module test harness in ``tests/lm_helpers.py``
    # and ``tests/test_matrix_excel.py``) load this file under a synthetic /
    # parentless package, so the package-relative ``...config`` cannot resolve.
    # ``app.config`` is stdlib-only, so load it directly by file path without
    # importing ``app/__init__`` (which would pull in Flask / SQLAlchemy).
    import importlib.util as _ilu
    import pathlib as _pathlib

    _config_path = _pathlib.Path(__file__).resolve().parents[2] / "config.py"
    _spec = _ilu.spec_from_file_location("_lm_pure_config", _config_path)
    _config_mod = _ilu.module_from_spec(_spec)
    _spec.loader.exec_module(_config_mod)
    _Config = _config_mod.Config

# --------------------------------------------------------------------------- #
# Schema (single source of truth, shared by import + export)
# --------------------------------------------------------------------------- #
# (attribute_key, japanese_header). Order == summary-sheet column order.
SUMMARY_COLUMNS: list[tuple[str, str]] = [
    ("test_id", "テストID"),            # calculated column (prefix + cat + no)
    ("category", "テスト区分"),
    ("category_name", "テスト区分名"),
    ("viewpoint", "テスト観点"),
    ("test_no", "テスト番号"),
    ("test_name", "テスト名"),
    ("purpose", "目的"),
    ("environment", "環境"),
    ("env_version", "環境バージョン"),
    ("data_flash", "データフラッシュ"),
    ("parameter", "パラメータ"),
    ("priority", "優先度"),
    ("description", "説明"),
    ("item_created", "項目作成"),
    ("exec_date", "実施日"),
    ("executor", "実施者"),
    ("version_label", "バージョン"),
    ("log", "ログ"),                    # calculated column
    ("result", "結果"),
    ("remark", "備考"),
    ("traceability_id", "トレーサビリティID"),
    ("upper_req_id", "上位要求ID"),
]
HEADER_TO_KEY = {jp: key for key, jp in SUMMARY_COLUMNS}
KEY_TO_HEADER = {key: jp for key, jp in SUMMARY_COLUMNS}
SUMMARY_HEADERS = [jp for _, jp in SUMMARY_COLUMNS]

# Columns that are calculated in the workbook and therefore not read on import.
CALCULATED_KEYS = {"test_id", "log"}
INT_KEYS = {"category", "test_no"}

# Metadata labels rendered (in order) on each detail-sheet block. Each maps to a
# summary column key so the VLOOKUP formula can target the right ``DB`` column.
DETAIL_LABELS: list[tuple[str, str]] = [
    ("目的", "purpose"),
    ("環境", "environment"),
    ("環境バージョン", "env_version"),
    ("データフラッシュ", "data_flash"),
    ("パラメータ", "parameter"),
    ("優先度", "priority"),
    ("説明", "description"),
    ("実施日", "exec_date"),
    ("実施者", "executor"),
    ("バージョン", "version_label"),
    ("ログ", "log"),
    ("結果", "result"),
    ("備考", "remark"),
]

# Defaults are configurable via ``.env`` (see app.config.Config.LM_TM_*); the
# literal fallbacks below preserve the historical workbook layout.
DEFAULT_SUMMARY_SHEET = _Config.LM_TM_SUMMARY_SHEET
DEFAULT_ID_PREFIX = _Config.LM_TM_ID_PREFIX
TABLE_NAME = "DB"

# Detail-sheet fixed column positions (1-based). Signals + timing follow F.
COL_STEP_NO = 2       # B 手順番号
COL_STEP_PURPOSE = 3  # C 手順目的
COL_STEP_OP = 4       # D 操作手順
COL_STEP_SUB = 5      # E サブルーチン
COL_STEP_ARG = 6      # F 引数
COL_SIGNAL_START = 7  # G .. first 入力値 signal

# Block layout offsets relative to the block header row.
_OFF_FIRST_LABEL = 1          # first metadata label row
_N_LABELS = len(DETAIL_LABELS)
_OFF_STEP_HEADER = _OFF_FIRST_LABEL + _N_LABELS + 1  # blank row before header
_OFF_SIGNAL1 = _OFF_STEP_HEADER + 1
_OFF_SIGNAL2 = _OFF_STEP_HEADER + 2
_OFF_FIRST_STEP = _OFF_STEP_HEADER + 3
_BLANK_ROWS_BETWEEN_BLOCKS = 2
_FIRST_BLOCK_ROW = 3          # rows 1 (A1 formula) + 2 (blank) precede it


class MatrixExcelError(Exception):
    """Raised when a workbook cannot be parsed as a test matrix."""


# --------------------------------------------------------------------------- #
# Import
# --------------------------------------------------------------------------- #
def _as_seekable(source: Union[str, "BinaryIO"]) -> Union[str, io.BytesIO]:
    """Return a source openpyxl/zipfile can seek on.

    A path is passed through unchanged. A file-like object is copied into an
    in-memory :class:`io.BytesIO` when it is not fully seekable — Werkzeug's
    upload streams are ``SpooledTemporaryFile`` objects which, on some Python
    versions, do not implement ``seekable`` and break the underlying ``zipfile``
    reader used by ``.xlsx`` parsing.
    """
    if isinstance(source, (str, bytes)) and not hasattr(source, "read"):
        return source  # filesystem path
    if isinstance(source, io.BytesIO):
        source.seek(0)
        return source

    # File-like: verify it is genuinely seekable, otherwise buffer it.
    seekable = False
    try:
        seekable = bool(source.seekable())  # type: ignore[attr-defined]
    except Exception:  # noqa: BLE001 - missing/broken seekable()
        seekable = False
    if seekable:
        try:
            source.seek(0)  # type: ignore[attr-defined]
            return source
        except Exception:  # noqa: BLE001
            pass

    try:
        try:
            source.seek(0)  # type: ignore[attr-defined]
        except Exception:  # noqa: BLE001
            pass
        data = source.read()  # type: ignore[attr-defined]
    except Exception as exc:  # noqa: BLE001
        raise MatrixExcelError(f"Could not read uploaded file: {exc}") from exc
    if isinstance(data, str):
        data = data.encode("utf-8")
    return io.BytesIO(data)


def parse_workbook(
    source: Union[str, "BinaryIO"], *, source_filename: str = ""
) -> dict[str, Any]:
    """Parse an ``.xlsx`` workbook into a plain matrix dict.

    Returns::

        {
          "name": str,                # workbook display name
          "source_filename": str,
          "summary_sheet": str,
          "id_prefix": str,
          "items": [ {<summary keys>, "steps": {...}}, ... ],
        }

    Raises :class:`MatrixExcelError` on structural problems (missing summary
    sheet / required headers).
    """
    try:
        wb = load_workbook(_as_seekable(source), data_only=True, read_only=False)
    except MatrixExcelError:
        raise
    except Exception as exc:  # noqa: BLE001
        raise MatrixExcelError(f"Could not open workbook: {exc}") from exc

    summary_ws, header_row = _find_summary_sheet(wb)
    col_index = _map_headers(summary_ws, header_row)

    items: list[dict[str, Any]] = []
    id_prefix = ""
    for r in range(header_row + 1, summary_ws.max_row + 1):
        row_vals = {
            key: summary_ws.cell(r, idx).value for key, idx in col_index.items()
        }
        if _row_is_empty(row_vals):
            continue
        item = _normalise_summary_row(row_vals)
        # Skip rows with no identity at all (category + test_no both missing).
        if item["category"] is None and item["test_no"] is None:
            continue
        # Detect the id prefix from the first literal テストID we see.
        raw_id = row_vals.get("test_id")
        if not id_prefix and isinstance(raw_id, str) and not raw_id.startswith("="):
            id_prefix = _prefix_from_test_id(raw_id)
        items.append(item)

    # Attach procedure blocks parsed from the category detail sheets, keyed by
    # (category, test_no) so it is independent of the id prefix.
    steps_by_key, prefix_from_detail = _parse_detail_sheets(wb, summary_ws.title)
    if not id_prefix:
        id_prefix = prefix_from_detail or DEFAULT_ID_PREFIX
    for item in items:
        key = (item["category"], item["test_no"])
        item["steps"] = steps_by_key.get(
            key, {"input_signals": [], "expected_signals": [], "steps": []}
        )

    name = source_filename or getattr(wb, "path", "") or summary_ws.title
    name = re.sub(r"\.xlsx$", "", str(name), flags=re.IGNORECASE)
    return {
        "name": name,
        "source_filename": source_filename,
        "summary_sheet": summary_ws.title,
        "id_prefix": id_prefix,
        "items": items,
    }


def _find_summary_sheet(wb) -> tuple[Any, int]:
    """Locate the summary worksheet and its header row (1-based)."""
    required = {"テスト区分", "テスト番号"}
    # 0) The sheet explicitly configured as the summary/設定 sheet, by name.
    for ws in wb.worksheets:
        if ws.title == DEFAULT_SUMMARY_SHEET:
            hr = _detect_header_row(ws, required)
            if hr:
                return ws, hr
    # 1) A sheet carrying the DB table.
    for ws in wb.worksheets:
        try:
            tables = ws.tables
        except Exception:  # noqa: BLE001 - read_only quirks
            tables = {}
        if TABLE_NAME in tables:
            hr = _detect_header_row(ws, required)
            if hr:
                return ws, hr
    # 2) Any sheet whose first rows contain the required headers.
    for ws in wb.worksheets:
        hr = _detect_header_row(ws, required)
        if hr:
            return ws, hr
    raise MatrixExcelError(
        "No summary sheet found: expected a sheet with テスト区分 / テスト番号 headers."
    )


def _detect_header_row(ws, required: set[str], scan: int = 5) -> Optional[int]:
    for r in range(1, min(scan, ws.max_row) + 1):
        values = {
            (c.value.strip() if isinstance(c.value, str) else c.value)
            for c in ws[r]
        }
        if required.issubset(values):
            return r
    return None


def _map_headers(ws, header_row: int) -> dict[str, int]:
    """Map summary attribute keys -> 1-based column indexes for this sheet."""
    col_index: dict[str, int] = {}
    for cell in ws[header_row]:
        val = cell.value
        if isinstance(val, str):
            key = HEADER_TO_KEY.get(val.strip())
            if key and key not in col_index:
                col_index[key] = cell.column
    missing = {"category", "test_no"} - set(col_index)
    if missing:
        raise MatrixExcelError(
            "Summary sheet is missing required columns: "
            + ", ".join(KEY_TO_HEADER[k] for k in sorted(missing))
        )
    return col_index


def _row_is_empty(row_vals: dict[str, Any]) -> bool:
    return all(v is None or v == "" for v in row_vals.values())


# Placeholder glyphs the VHILS workbook writes into "not applicable" cells when
# 項目作成 (item_created) is 不要 — chiefly 実施日 (exec_date) and 実施者 (executor).
# These must be treated as empty, otherwise a lone dash breaks date coercion /
# field validation on import ("格式不正确").
_DASH_PLACEHOLDERS = frozenset({"-", "－", "—", "ー", "―", "‐", "–", "ｰ"})


def _is_dash_placeholder(val: Any) -> bool:
    return isinstance(val, str) and val.strip() in _DASH_PLACEHOLDERS


def _normalise_summary_row(row_vals: dict[str, Any]) -> dict[str, Any]:
    item: dict[str, Any] = {}
    for key, _jp in SUMMARY_COLUMNS:
        # ``test_id`` / ``log`` are Excel-calculated columns: read back with
        # ``data_only=True`` they carry the cached result (e.g. "VH001001"), so
        # importing them preserves the workbook's テストID / ログ content instead
        # of dropping it. On export they are regenerated as formulas regardless.
        val = row_vals.get(key)
        # A bare dash is a "not applicable" marker, not real data — drop it so
        # date/number coercion and validation don't reject the whole row.
        if _is_dash_placeholder(val):
            val = None
        if key in INT_KEYS:
            item[key] = _to_int(val)
        elif key == "exec_date":
            item[key] = _to_date_str(val)
        else:
            item[key] = "" if val is None else str(val)
    return item


def _to_int(val: Any) -> Optional[int]:
    if val is None or val == "":
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        m = re.search(r"-?\d+", str(val))
        return int(m.group()) if m else None


def _to_date_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, (_dt.datetime, _dt.date)):
        return val.isoformat()
    return str(val)


def _prefix_from_test_id(test_id: str) -> str:
    """Strip the trailing 6 identity digits (cat3 + no3) to recover the prefix."""
    if re.search(r"\d{6}$", test_id):
        return test_id[:-6]
    return test_id


def _is_detail_sheet_name(title: Any) -> bool:
    """A category detail sheet is named after its numeric テスト区分 (e.g. "1",
    "12"). Only such pure-numeric sheets are parsed for procedure blocks; any
    other sheet (cover pages, notes, revision history, etc.) is ignored so it
    can never abort the import."""
    return str(title).strip().isdigit()


def _parse_detail_sheets(
    wb, summary_title: str
) -> tuple[dict[tuple[Optional[int], Optional[int]], dict], str]:
    """Parse only the numeric category detail sheets (plus skip the summary
    sheet). Non-numeric / auxiliary sheets are ignored, and any malformed detail
    sheet is skipped rather than failing the whole import."""
    result: dict[tuple[Optional[int], Optional[int]], dict] = {}
    prefix = ""
    for ws in wb.worksheets:
        if ws.title == summary_title:
            continue
        if not _is_detail_sheet_name(ws.title):
            continue
        try:
            blocks = list(_iter_detail_blocks(ws))
        except Exception:  # noqa: BLE001 - a bad detail sheet must not abort import
            continue
        for block in blocks:
            test_id = block["test_id"]
            if not prefix and isinstance(test_id, str):
                prefix = _prefix_from_test_id(test_id)
            key = _key_from_test_id(test_id)
            if key is not None:
                result[key] = {
                    "input_signals": block["input_signals"],
                    "expected_signals": block["expected_signals"],
                    "steps": block["steps"],
                }
    return result, prefix


def _key_from_test_id(test_id: Any) -> Optional[tuple[int, int]]:
    if not isinstance(test_id, str):
        return None
    m = re.search(r"(\d{3})(\d{3})$", test_id)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def _iter_detail_blocks(ws):
    """Yield one dict per procedure block on a detail worksheet."""
    max_row = ws.max_row
    r = 1
    while r <= max_row:
        b = ws.cell(r, COL_STEP_NO).value
        if isinstance(b, str) and b.startswith("ID;;"):
            yield _parse_block(ws, r)
            # Skip past this block's steps to avoid re-scanning.
            r = _block_end_row(ws, r) + 1
            continue
        r += 1


def _parse_block(ws, header_row: int) -> dict[str, Any]:
    test_id = ws.cell(header_row, COL_STEP_NO).value
    step_header_row = _find_step_header(ws, header_row)
    input_signals: list[list[str]] = []
    expected_signals: list[list[str]] = []
    steps: list[dict[str, Any]] = []
    if step_header_row is not None:
        input_signals, expected_signals, steps = parse_step_table(ws, step_header_row)
    return {
        "test_id": test_id,
        "input_signals": input_signals,
        "expected_signals": expected_signals,
        "steps": steps,
    }


def parse_step_table(ws, step_header_row: int, *, stop_on_block=None):
    """Parse one procedure table into (input_signals, expected_signals, steps).

    The procedure-table geometry (``手順番号 .. 確認タイミング`` header on
    ``step_header_row``, two signal-name header rows, then the step rows) is
    shared verbatim by the Test-Matrix detail sheets and the Lib(Func) workbook,
    so both parsers call this. ``stop_on_block`` is an optional
    ``callable(ws, row) -> bool`` used by the Lib parser to stop at the next
    function block header (Test-Matrix blocks are delimited by the ``ID;;``
    marker, which is always checked).
    """
    n_in, n_exp, timing_col = _measure_step_columns(ws, step_header_row)
    sig1 = step_header_row + 1
    sig2 = step_header_row + 2
    input_signals: list[list[str]] = []
    expected_signals: list[list[str]] = []
    for i in range(n_in):
        c = COL_SIGNAL_START + i
        input_signals.append([_txt(ws.cell(sig1, c).value), _txt(ws.cell(sig2, c).value)])
    for i in range(n_exp):
        c = COL_SIGNAL_START + n_in + i
        expected_signals.append([_txt(ws.cell(sig1, c).value), _txt(ws.cell(sig2, c).value)])
    steps: list[dict[str, Any]] = []
    r = step_header_row + 3
    while r <= ws.max_row:
        no = ws.cell(r, COL_STEP_NO).value
        if no is None or (isinstance(no, str) and no.startswith("ID;;")):
            break
        if stop_on_block is not None and stop_on_block(ws, r):
            break
        steps.append(
            {
                "no": _to_int(no) if _to_int(no) is not None else no,
                "purpose": _cell_or_none(ws.cell(r, COL_STEP_PURPOSE).value),
                "operation": _cell_or_none(ws.cell(r, COL_STEP_OP).value),
                "subroutine": _cell_or_none(ws.cell(r, COL_STEP_SUB).value),
                "args": _cell_or_none(ws.cell(r, COL_STEP_ARG).value),
                "inputs": [
                    _cell_or_none(ws.cell(r, COL_SIGNAL_START + i).value)
                    for i in range(n_in)
                ],
                "expecteds": [
                    _cell_or_none(ws.cell(r, COL_SIGNAL_START + n_in + i).value)
                    for i in range(n_exp)
                ],
                "timing": _cell_or_none(ws.cell(r, timing_col).value),
            }
        )
        r += 1
    return input_signals, expected_signals, steps


def _find_step_header(ws, header_row: int, look: int = 40) -> Optional[int]:
    for r in range(header_row + 1, min(header_row + look, ws.max_row) + 1):
        if ws.cell(r, COL_STEP_NO).value == "手順番号":
            return r
    return None


def _measure_step_columns(ws, step_header_row: int) -> tuple[int, int, int]:
    """Return (n_input_signals, n_expected_signals, timing_column_index)."""
    n_in = n_exp = 0
    timing_col = COL_SIGNAL_START
    c = COL_SIGNAL_START
    while c <= ws.max_column:
        v = ws.cell(step_header_row, c).value
        if v == "入力値":
            n_in += 1
        elif v == "期待値":
            n_exp += 1
        elif v == "確認タイミング":
            timing_col = c
            break
        elif v is None:
            break
        c += 1
    if timing_col <= COL_SIGNAL_START + n_in + n_exp - 1:
        timing_col = COL_SIGNAL_START + n_in + n_exp
    return n_in, n_exp, timing_col


def _block_end_row(ws, header_row: int) -> int:
    step_header_row = _find_step_header(ws, header_row)
    if step_header_row is None:
        return header_row + _OFF_FIRST_STEP
    r = step_header_row + 3
    last = r - 1
    while r <= ws.max_row:
        no = ws.cell(r, COL_STEP_NO).value
        if no is None or (isinstance(no, str) and no.startswith("ID;;")):
            break
        last = r
        r += 1
    return last


def _txt(val: Any) -> str:
    return "" if val is None else str(val)


def _cell_or_none(val: Any) -> Any:
    if val is None:
        return None
    if isinstance(val, (_dt.datetime, _dt.date)):
        return val.isoformat()
    return val


# --------------------------------------------------------------------------- #
# Export
# --------------------------------------------------------------------------- #
def build_workbook(matrix: dict[str, Any]) -> Workbook:
    """Regenerate an ``.xlsx`` workbook from a matrix dict (see parse_workbook).

    The returned :class:`openpyxl.Workbook` has the summary ``DB`` table sheet
    plus one detail sheet per category, faithful to the source layout.
    """
    id_prefix = matrix.get("id_prefix") or DEFAULT_ID_PREFIX
    summary_title = matrix.get("summary_sheet") or DEFAULT_SUMMARY_SHEET
    items = matrix.get("items") or []

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = summary_title
    _write_summary_sheet(summary_ws, items, id_prefix)

    # One detail sheet per category (only items that carry a procedure block).
    by_cat: dict[int, list[dict]] = {}
    for it in items:
        cat = it.get("category")
        steps = (it.get("steps") or {}).get("steps") or []
        if cat is None or not steps:
            continue
        it = {**it, "_id_prefix": id_prefix}
        by_cat.setdefault(cat, []).append(it)
    for cat in sorted(by_cat):
        ws = wb.create_sheet(title=str(cat))
        _write_detail_sheet(ws, by_cat[cat])

    return wb


def _write_summary_sheet(ws, items: list[dict], id_prefix: str) -> None:
    ws.append(SUMMARY_HEADERS)
    for i, it in enumerate(items):
        r = i + 2
        for key, jp in SUMMARY_COLUMNS:
            col = SUMMARY_HEADERS.index(jp) + 1
            cell = ws.cell(r, col)
            if key == "test_id":
                cell.value = f'="{id_prefix}"&TEXT(B{r},"000")&TEXT(E{r},"000")'
            elif key == "log":
                # =IF(優先度="高", テストID, "-")
                cell.value = f'=IF(L{r}="高",A{r},"-")'
            elif key == "exec_date":
                cell.value = _date_cell(it.get("exec_date"))
            else:
                val = it.get(key)
                cell.value = None if val in (None, "") else val

    last_col = get_column_letter(len(SUMMARY_HEADERS))
    last_row = len(items) + 1
    ref = f"A1:{last_col}{last_row}"
    table = Table(displayName=TABLE_NAME, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    _apply_summary_widths(ws)


_SUMMARY_WIDTHS = {
    "A": 34, "B": 9.5, "C": 17, "D": 12, "E": 9.5, "F": 68, "G": 88, "H": 7.5,
    "I": 12, "J": 13, "K": 35, "L": 8.8, "M": 28.5, "N": 10, "O": 10, "P": 8.8,
    "Q": 10, "R": 34, "S": 7.5, "T": 10, "U": 20.5, "V": 32,
}


def _apply_summary_widths(ws) -> None:
    for letter, width in _SUMMARY_WIDTHS.items():
        ws.column_dimensions[letter].width = width


# A1 category lookup: derives the category number from the sheet (tab) name.
_A1_CATEGORY_FORMULA = (
    '=VLOOKUP(INT(RIGHT(CELL("filename",A2),LEN(CELL("filename",A2))'
    '-FIND("]",CELL("filename",A2)))),DB[[テスト区分]:[テスト区分名]],'
    'COLUMN(DB[テスト区分名])-COLUMN(DB[テスト区分])+1,0)'
)


def _write_detail_sheet(ws, items: list[dict]) -> None:
    ws.cell(1, 1).value = _A1_CATEGORY_FORMULA
    ws.freeze_panes = "E1"
    _apply_detail_widths(ws)

    row = _FIRST_BLOCK_ROW
    for it in items:
        row = _write_detail_block(ws, row, it)
        row += _BLANK_ROWS_BETWEEN_BLOCKS


def _write_detail_block(ws, header_row: int, item: dict) -> int:
    """Write one procedure block starting at ``header_row``; return last used row."""
    test_id = _reconstruct_test_id(item)
    hr = header_row
    ws.cell(hr, COL_STEP_NO).value = test_id                       # B: test id
    ws.cell(hr, COL_STEP_OP).value = (                             # D: テスト名 VLOOKUP
        f"=VLOOKUP(B{hr},DB[#Data],COLUMN(DB[テスト名]),0)"
    )

    # Metadata label rows (VLOOKUP into DB, referencing this block's test id).
    for j, (label, key) in enumerate(DETAIL_LABELS):
        rr = hr + _OFF_FIRST_LABEL + j
        ws.cell(rr, COL_STEP_NO).value = label
        col_header = KEY_TO_HEADER[key]
        if key == "exec_date":
            ws.cell(rr, COL_STEP_PURPOSE).value = (
                f'=": "&TEXT(VLOOKUP(B{hr},DB[#Data],'
                f'COLUMN(DB[{col_header}]),0),"yyyy/mm/dd")'
            )
        else:
            ws.cell(rr, COL_STEP_PURPOSE).value = (
                f'=": "&VLOOKUP(B{hr},DB[#Data],COLUMN(DB[{col_header}]),0)'
            )

    steps = item.get("steps") or {}
    input_signals = steps.get("input_signals") or []
    expected_signals = steps.get("expected_signals") or []
    step_rows = steps.get("steps") or []
    n_in = len(input_signals)
    n_exp = len(expected_signals)
    timing_col = COL_SIGNAL_START + n_in + n_exp

    # Procedure table header.
    sh = hr + _OFF_STEP_HEADER
    ws.cell(sh, COL_STEP_NO).value = "手順番号"
    ws.cell(sh, COL_STEP_PURPOSE).value = "手順目的"
    ws.cell(sh, COL_STEP_OP).value = "操作手順"
    ws.cell(sh, COL_STEP_SUB).value = "サブルーチン"
    ws.cell(sh, COL_STEP_ARG).value = "引数"
    for i in range(n_in):
        ws.cell(sh, COL_SIGNAL_START + i).value = "入力値"
    for i in range(n_exp):
        ws.cell(sh, COL_SIGNAL_START + n_in + i).value = "期待値"
    ws.cell(sh, timing_col).value = "確認タイミング"

    # Signal-name header rows.
    s1 = hr + _OFF_SIGNAL1
    s2 = hr + _OFF_SIGNAL2
    for i, (name, path) in enumerate(input_signals):
        ws.cell(s1, COL_SIGNAL_START + i).value = name or None
        ws.cell(s2, COL_SIGNAL_START + i).value = path or None
    for i, (name, path) in enumerate(expected_signals):
        ws.cell(s1, COL_SIGNAL_START + n_in + i).value = name or None
        ws.cell(s2, COL_SIGNAL_START + n_in + i).value = path or None

    # Step rows.
    last_row = s2
    for k, st in enumerate(step_rows):
        rr = hr + _OFF_FIRST_STEP + k
        last_row = rr
        ws.cell(rr, COL_STEP_NO).value = st.get("no")
        ws.cell(rr, COL_STEP_PURPOSE).value = st.get("purpose")
        ws.cell(rr, COL_STEP_OP).value = st.get("operation")
        ws.cell(rr, COL_STEP_SUB).value = st.get("subroutine")
        ws.cell(rr, COL_STEP_ARG).value = st.get("args")
        inputs = st.get("inputs") or []
        expecteds = st.get("expecteds") or []
        for i in range(n_in):
            ws.cell(rr, COL_SIGNAL_START + i).value = (
                inputs[i] if i < len(inputs) else None
            )
        for i in range(n_exp):
            ws.cell(rr, COL_SIGNAL_START + n_in + i).value = (
                expecteds[i] if i < len(expecteds) else None
            )
        ws.cell(rr, timing_col).value = st.get("timing")
    return last_row


_DETAIL_WIDTHS = {"A": 3, "B": 9, "C": 26.5, "D": 47.8, "E": 33.6, "F": 17.2}


def _apply_detail_widths(ws) -> None:
    for letter, width in _DETAIL_WIDTHS.items():
        ws.column_dimensions[letter].width = width


def _reconstruct_test_id(item: dict) -> str:
    tid = item.get("test_id")
    if isinstance(tid, str) and tid and not tid.startswith("="):
        return tid
    prefix = item.get("_id_prefix") or ""
    cat = item.get("category")
    no = item.get("test_no")
    if cat is not None and no is not None:
        return f"{prefix}{cat:03d}{no:03d}"
    return prefix


def _date_cell(value: Any) -> Any:
    """Turn a stored exec_date string back into a datetime (or pass through)."""
    if value in (None, ""):
        return None
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value
    text = str(value)
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return _dt.datetime.strptime(text, fmt)
        except ValueError:
            continue
    return text
